import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

name = []
# program read information from people.csv file and put all data in name list.
with open("people.csv", "r") as f:
    next(f)
    for l in f:
        row = l.rstrip().split(",")
        vards = row[2]
        uzvards = row[3]
        name.append(f"{vards} {uzvards}")


url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(2)

code_fullname = []
for i in range(len(name)):
    fullname = name[i]

    atrast = driver.find_element(By.ID, "input")
    atrast.click()
    atrast.clear()
    atrast.send_keys(fullname)

    result_code = driver.find_element(By.ID, "output")
    name_code = result_code.get_attribute('value')

    code_fullname.append((fullname.strip(), name_code.strip()))
driver.quit()

wb = load_workbook('salary.xlsx')
ws = wb.active
vards_alga_dict = {}

for row in range(2, ws.max_row + 1):
    a = ws[f'A{row}'].value  # vards codets
    b = ws[f'B{row}'].value  # alga

    if a is not None:
        matching_names = [name[0] for name in code_fullname if a.strip() == name[1].strip()]

        if matching_names:
            vards_alga_dict[matching_names[0]] = vards_alga_dict.get(matching_names[0], 0) + b

user_input = input("Full name: ").strip()
matching_salary = vards_alga_dict.get(user_input, None)

if matching_salary is not None:
    print(f"{matching_salary}")
else:
    print("Name not found")
